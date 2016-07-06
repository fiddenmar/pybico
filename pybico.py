import getopt
import sys
import string
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pymysql.cursors

import regex as re

PYBICO_VERBOSE = False

def split_authors(res):
	authors = res.split(', ')
	new_authors = []
	for author in authors:
		new_authors.append(author.strip().replace(". ", "."))
	return new_authors

def isEnglish(s):
    try:
        s.encode('ascii')
    except UnicodeEncodeError:
        return False
    else:
        return True

def pybico_import_txt(filename):
	rn1 = "(?P<authors>((\pL\. ?(\pL\. )?\pL+,? )|(\pL+ \pL\. ?(\pL\.)?,? )" #regular for authors
	rn2 = "|(\p{Lu}\p{Ll}+ \p{Lu}\p{Ll}+,? )"
	rn3 = ")+)"
	ra_ru = "(?P<article>\p{Lu}\p{Ll}+ \p{Ll}+.*?) *\/\/ *" #regular for article
	ra_eng = "(?P<article>\p{Lu}.*?) *\/\/ *" #regular for article
	rj = '(?P<source>[ \pL"“”]+)' #regular for source
	rm = "(?P<misc>.+)" #regular for misc
	reg_ru = re.compile(rn1+rn2+rn3+ra_ru+rj+rm, re.UNICODE)
	reg_eng = re.compile(rn1+rn3+ra_eng+rj+rm, re.UNICODE)
	data = []
	f = open(filename, 'r')
	content = f.read()
	items = content.split('\n')
	for item in items:
		res = None
		if isEnglish(item[:15]):
			res = reg_eng.match(item.strip())
		else:
			res = reg_ru.match(item.strip())
		if res != None:
			authors = res.group("authors").split(', ')
			data.append({"authors": split_authors(res.group("authors")), "article": res.group("article"), "source": res.group("source"), "misc": res.group("misc")})
		else:
			print("Wrong line: " + item)	
	return data

def pybico_import(format, filename, usr, pswd):
	data = []
	#test = ([author1, author2, ...], article, source, misc)
	if format == "txt":
		data = pybico_import_txt(filename)
	else:
		data = None

	connection = pymysql.connect(host="127.0.0.1",
								user=usr,
								password=pswd,
								db='pybico',
								charset='utf8',
								cursorclass=pymysql.cursors.DictCursor)
	try:
		for item in data:
			author_id = []
			source_id = 0
			publication_id = 0
			with connection.cursor() as cursor:
				cursor.execute('SET NAMES utf8;')
				cursor.execute('SET CHARACTER SET utf8;')
				cursor.execute('SET character_set_connection=utf8;')
				get_source_sql = "SELECT `id` FROM `source` WHERE `name`=%s"
				cursor.execute(get_source_sql, (item["source"]))
				result = cursor.fetchone()
				if result:
					source_id = result["id"]
				else:
					insert_source_sql = "INSERT INTO `source` (`name`) VALUES (%s)"
					cursor.execute(insert_source_sql, (item["source"]))
					connection.commit()
					cursor.execute(get_source_sql, (item["source"]))
					res = cursor.fetchone()
					source_id = res["id"]
			for a in item["authors"]:
				author = a.strip()
				if author != "":
					with connection.cursor() as cursor:
						get_author_sql = "SELECT `id` FROM `author` WHERE `name`=%s"
						cursor.execute(get_author_sql, (author))
						result = cursor.fetchone()
						if result:
							author_id.append(result["id"])
						else:
							insert_author_sql = "INSERT INTO `author` (`name`) VALUES (%s)"
							cursor.execute(insert_author_sql, (author))
							connection.commit()
							cursor.execute(get_author_sql, (author))
							res = cursor.fetchone()
							author_id.append(res["id"])
			with connection.cursor() as cursor:
				get_publication_sql = "SELECT `id` FROM `publication` WHERE `title`=%s"
				cursor.execute(get_publication_sql, (item["article"]))
				result = cursor.fetchone()
				if result:
					publication_id = result["id"]
				else:
					insert_publication_sql = "INSERT INTO `publication` (`title`, `source_id`, `misc`) VALUES (%s, %s, %s)"
					cursor.execute(insert_publication_sql, (item["article"], str(source_id), item["misc"]))
					connection.commit()
					cursor.execute(get_publication_sql, (item["article"]))
					res = cursor.fetchone()
					publication_id = res["id"]
					for author in author_id:
						insert_relation_sql = "INSERT INTO `relation` (`publication_id`, `author_id`) VALUES (%s, %s)"
						cursor.execute(insert_relation_sql, (str(publication_id), str(author)))
					connection.commit()
	finally:
		connection.close()
	# return data

def pos_to_char(i):
	return string.ascii_uppercase[i] 

def pybico_extract(usr, pswd):
	data = []
	connection = pymysql.connect(host="127.0.0.1",
								user=usr,
								password=pswd,
								db='pybico',
								charset='utf8',
								cursorclass=pymysql.cursors.DictCursor)
	try:
		with connection.cursor() as cursor:
			get_publication_sql = "SELECT * FROM `publication` WHERE `registered`=%s"
			cursor.execute(get_publication_sql, (str(0)))
			result = cursor.fetchall()
			for pub in result:
				target_id = pub["id"]
				authors = []
				source = ""
				get_authors_sql = "SELECT `author_id` FROM `relation` WHERE `publication_id`=%s"
				cursor.execute(get_authors_sql, (str(target_id)))
				res = cursor.fetchall()
				for a in res:
					get_name_sql = "SELECT `name`, `miet` FROM `author` WHERE `id`=%s"
					cursor.execute(get_name_sql, (str(a["author_id"])))
					author = cursor.fetchone()
					authors.append({"name": author["name"], "miet": author["miet"]})
				get_source_sql = "SELECT `name`, `type`, `scopus`, `wos`, `hac`, `rsci` FROM `source` WHERE `id`=%s"
				cursor.execute(get_source_sql, (str(pub["source_id"])))
				src_res = cursor.fetchone()
				source = {"name": src_res["name"], "type": src_res["type"], "scopus": src_res["scopus"], "wos": src_res["wos"], "hac": src_res["hac"], "rsci": src_res["rsci"]}
				data.append({"authors": authors, "article": pub["title"], "source": source, "misc": pub["misc"]})
	finally:
		connection.close()
	return data

def get_author_names(authors):
	names = []
	for author in authors:
		names.append(author["name"])
	return names

def get_author_miet(authors):
	miet = 0
	for author in authors:
		miet += not not author["miet"]
	return miet

def get_impact_factor(source):
	impact_list = [source["scopus"], source["wos"], source["hac"], source["rsci"]]
	impact = max(impact_list)
	if impact == 0:
		impact = -1
	if impact == -1:
		return ""
	return impact

def get_status(source):
	status = []
	impact_list = [source["scopus"], source["wos"], source["hac"], source["rsci"]]
	status_list = ["Scopus", "Web of Science", "ВАК", "РИНЦ"]
	for i, impact in enumerate(impact_list):
		if impact != 0:
			status.append(status_list[i])
	return status

def pybico_export_xlsx(filename, data):
	offset = 3
	wb = Workbook()
	ws = wb.active
	wstyle1 = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
	wstyle2 = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
	for r in range(0, offset+len(data)+1):
		ws.row_dimensions[r+1].height = 100
	ws.row_dimensions[3].height = 15
	for c in range(1,9):
		ch = pos_to_char(c-1)
		ws.column_dimensions[ch].width = 20
	for r in range(1,offset+1):
		for c in range(1,10):
			ws.cell(row = r, column = c).alignment = wstyle2
	for r in range(offset+1, offset+len(data)+1):
		for c in (2,3,4,5):
			ws.cell(row = r, column = c).alignment = wstyle1
		for c in (1,6,7,8,9):
			ws.cell(row = r, column = c).alignment = wstyle2
	ws.merge_cells('H1:I1')
	ws['A1'] = "№ раздела"
	ws['B1'] = "Автор (ФИО сотрудника МИЭТ, студента, аспиранта)"
	ws['C1'] = "Название статьи, книги, монографии, уч. пособия и др."
	ws['D1'] = "Наименование журнала или конференции"
	ws['E1'] = "Город, издательство, год, номер, том, страницы"
	ws['F1'] = "Статус\nWeb of Science\nScopus\nРИНЦ\nВАК"
	ws['G1'] = "Импакт-фактор"
	ws['H1'] = "Количество авторов"
	ws['H2'] = "Всего"
	ws['I2'] = "В т.ч. сотрудников МИЭТ"
	for c in range(1,9):
		ws.cell(row = 3, column = c).value = c
	for i, item in enumerate(data):
		pos = str(i+1+offset)
		ws['A'+pos] = item["source"]["type"]
		ws['B'+pos] = "\n".join(get_author_names(item["authors"]))
		ws['C'+pos] = item["article"]
		ws['D'+pos] = item["source"]["name"]
		ws['E'+pos] = item["misc"]
		ws['F'+pos] = str("\n".join(get_status(item["source"])))
		ws['G'+pos] = str(get_impact_factor(item["source"]))
		ws['H'+pos] = str(len(item["authors"]))
		ws['I'+pos] = str(get_author_miet(item["authors"]))

	wb.save(filename) 

def pybico_export(format, filename, data):
	if format == "xlsx":
		pybico_export_xlsx(filename, data)
	else:
		return None
	
def usage():
	print("usage: pybico [options]")
	print("\toptions:")
	print("\t -h, --help\t print out help")
	print("\t -v\t verbose mode")
	print("\t -i\t path to import file")
	print("\t -e\t path to export file")
	print("\t --if\t import format (txt)")
	print("\t --ef\t export format (xlsx)")
	print("\t -u\t database user")
	print("\t -p\t path to database password")

def main(argv):
	global PYBICO_VERBOSE

	try:
		opts, args = getopt.getopt(argv, "hvi:e:u:p:", ["help", "if", "ef"])
	except getopt.GetoptError as err:
		print(str(err))
		usage()
		sys.exit(2)

	PYBICO_VERBOSE = False
	import_format = "txt"
	export_format = "xlsx"
	import_filename = ""
	export_filename = ""
	password_path = ""
	user = ""

	for o, a in opts:
		if o == "-v":
			PYBICO_VERBOSE = True
		elif o in ("-h", "--help"):
			usage()
			sys.exit()
		elif o == "-u":
			user = a
		elif o == "-p":
			password_path = a
		elif o == "-i":
			import_filename = a
		elif o == "-e":
			export_filename = a
		elif o == "--if":
			import_format = a
		elif o == "--ef":
			export_format = a
		else:
			assert False, "unhandled option"

	f = open(password_path, 'r')
	password = f.read()

	if import_filename != "":
		pybico_import(import_format, import_filename, user, password)
	if export_filename != "":
		data = pybico_extract(user, password)
		pybico_export(export_format, export_filename, data)

if __name__ == '__main__':
	main(sys.argv[1:])